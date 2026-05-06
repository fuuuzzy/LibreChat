import io
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

_TZ_UTC8 = timezone(timedelta(hours=8))


def _to_utc8(dt):
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_TZ_UTC8).strftime("%Y-%m-%d %H:%M:%S")
    return str(dt)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("utf-8")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def export_transactions(transactions: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Token对账"

    headers = [
        "时间", "用户", "邮箱", "模型", "类型",
        "Token数", "会话ID",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for tx in transactions:
        user_info = tx.get("userInfo") or {}
        token_type = tx.get("tokenType", "")
        type_label = "输入" if token_type == "prompt" else ("输出" if token_type == "completion" else token_type)
        ws.append([
            _to_utc8(tx.get("createdAt")),
            user_info.get("name", ""),
            user_info.get("email", ""),
            tx.get("model", ""),
            type_label,
            abs(tx.get("tokenValue", 0) or 0),
            tx.get("conversationId", ""),
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_messages(messages: list[dict], conversation_title: str = "") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "会话流水"[:31]

    headers = [
        "时间", "发送者", "模型", "内容", "Token数",
        "是否用户消息", "是否有错误", "消息ID",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for msg in messages:
        text = msg.get("text", "") or ""
        # Truncate very long messages for Excel
        if len(text) > 5000:
            text = text[:5000] + "...(截断)"
        ws.append([
            _to_utc8(msg.get("createdAt")),
            msg.get("sender", ""),
            msg.get("model", ""),
            text,
            msg.get("tokenCount", 0),
            "是" if msg.get("isCreatedByUser") else "否",
            "是" if msg.get("error") else "否",
            msg.get("messageId", ""),
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_usage_records(records: list[dict]) -> io.BytesIO:
    """Export usage records (merged by messageId) to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "使用记录明细"

    headers = [
        "时间", "用户", "邮箱", "模型", "来源", "输入Token", "输出Token",
        "缓存写入Token", "缓存读取Token",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for rec in records:
        user_info = rec.get("userInfo") or {}
        ws.append([
            _to_utc8(rec.get("createdAt")),
            user_info.get("name", ""),
            user_info.get("email", ""),
            rec.get("model", ""),
            rec.get("context", ""),
            rec.get("promptTokens", 0),
            rec.get("completionTokens", 0),
            rec.get("writeTokens", 0),
            rec.get("readTokens", 0),
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
