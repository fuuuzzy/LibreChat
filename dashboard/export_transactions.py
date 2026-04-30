#!/usr/bin/env python3
"""Export LibreChat transaction/consumption records from MongoDB to Excel."""

import argparse
import sys
from datetime import datetime, timedelta, timezone

_TZ_UTC8 = timezone(timedelta(hours=8))

try:
    from pymongo import MongoClient
except ImportError:
    sys.exit("Missing dependency: pip install pymongo")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


def parse_args():
    parser = argparse.ArgumentParser(description="Export LibreChat transactions to Excel")
    parser.add_argument(
        "-c", "--connection",
        required=True,
        help='MongoDB connection string, e.g. "mongodb://user:pass@host:27017/librechat"',
    )
    parser.add_argument(
        "-s", "--start",
        required=True,
        help="Start date (YYYY-MM-DD), inclusive",
    )
    parser.add_argument(
        "-e", "--end",
        required=True,
        help="End date (YYYY-MM-DD), inclusive",
    )
    parser.add_argument(
        "-o", "--output",
        default="transactions_export.xlsx",
        help="Output Excel file path (default: transactions_export.xlsx)",
    )
    return parser.parse_args()


def parse_date(date_str: str) -> datetime:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.replace(tzinfo=_TZ_UTC8).astimezone(timezone.utc)


def fetch_users(db) -> dict[str, dict]:
    """Build a mapping: user ObjectId string -> {name, email}."""
    users = {}
    for u in db.users.find({}, {"_id": 1, "name": 1, "username": 1, "email": 1}):
        uid = str(u["_id"])
        users[uid] = {
            "name": u.get("name") or u.get("username") or "",
            "email": u.get("email") or "",
        }
    return users


def fetch_transactions(db, start: datetime, end: datetime) -> list[dict]:
    """Fetch transactions within [start, end+1day)."""
    end_exclusive = end.replace(day=end.day + 1) if end.day < 28 else end.replace(
        month=end.month + 1 if end.month < 12 else 1,
        day=1,
        year=end.year + 1 if end.month == 12 else end.year,
    )
    # Safest: use $lt with next day at 00:00
    from datetime import timedelta
    end_exclusive = end + timedelta(days=1)

    query = {
        "createdAt": {
            "$gte": start,
            "$lt": end_exclusive,
        }
    }
    return list(db.transactions.find(query).sort("createdAt", 1))


def compute_token_cost(doc: dict) -> float:
    """Compute USD cost for a transaction record.

    Balance schema comment: 1000 tokenCredits = 1 mill ($0.001 USD)
    tokenValue is already in tokenCredits. rawAmount may also be present.
    """
    token_value = doc.get("tokenValue") or 0
    # tokenCredits -> USD: divide by 1_000_000
    return token_value / 1_000_000 if token_value else 0.0


def write_raw_sheet(wb: Workbook, transactions: list[dict], users: dict):
    ws = wb.active
    ws.title = "消费记录"

    headers = [
        "时间", "用户", "邮箱", "模型", "Token类型",
        "Input Tokens", "Write Tokens", "Read Tokens",
        "Token Value (credits)", "费用 (USD)", "对话ID", "消息ID",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, doc in enumerate(transactions, 2):
        uid = str(doc.get("user", ""))
        user_info = users.get(uid, {"name": uid, "email": ""})
        cost = compute_token_cost(doc)

        created_at = doc.get("createdAt")
        if isinstance(created_at, datetime):
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            created_at = created_at.astimezone(_TZ_UTC8).strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=i, column=1, value=created_at or "")
        ws.cell(row=i, column=2, value=user_info["name"])
        ws.cell(row=i, column=3, value=user_info["email"])
        ws.cell(row=i, column=4, value=doc.get("model", ""))
        ws.cell(row=i, column=5, value=doc.get("tokenType", ""))
        ws.cell(row=i, column=6, value=doc.get("inputTokens") or 0)
        ws.cell(row=i, column=7, value=doc.get("writeTokens") or 0)
        ws.cell(row=i, column=8, value=doc.get("readTokens") or 0)
        ws.cell(row=i, column=9, value=doc.get("tokenValue") or 0)
        ws.cell(row=i, column=10, value=cost)
        ws.cell(row=i, column=10).number_format = '#,##0.000000'
        ws.cell(row=i, column=11, value=doc.get("conversationId", ""))
        ws.cell(row=i, column=12, value=doc.get("messageId", ""))

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def write_summary_sheet(wb: Workbook, transactions: list[dict], users: dict):
    ws = wb.create_sheet(title="用户汇总")

    headers = [
        "用户", "邮箱", "总 Input Tokens", "总 Write Tokens", "总 Read Tokens",
        "总 Token Value (credits)", "总费用 (USD)", "调用次数",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Aggregate by user
    user_stats: dict[str, dict] = {}
    for doc in transactions:
        uid = str(doc.get("user", ""))
        if uid not in user_stats:
            user_stats[uid] = {
                "inputTokens": 0,
                "writeTokens": 0,
                "readTokens": 0,
                "tokenValue": 0,
                "cost": 0.0,
                "count": 0,
            }
        s = user_stats[uid]
        s["inputTokens"] += doc.get("inputTokens") or 0
        s["writeTokens"] += doc.get("writeTokens") or 0
        s["readTokens"] += doc.get("readTokens") or 0
        s["tokenValue"] += doc.get("tokenValue") or 0
        s["cost"] += compute_token_cost(doc)
        s["count"] += 1

    # Sort by cost descending
    sorted_users = sorted(user_stats.items(), key=lambda x: x[1]["cost"], reverse=True)

    for i, (uid, stats) in enumerate(sorted_users, 2):
        user_info = users.get(uid, {"name": uid, "email": ""})
        ws.cell(row=i, column=1, value=user_info["name"])
        ws.cell(row=i, column=2, value=user_info["email"])
        ws.cell(row=i, column=3, value=stats["inputTokens"])
        ws.cell(row=i, column=4, value=stats["writeTokens"])
        ws.cell(row=i, column=5, value=stats["readTokens"])
        ws.cell(row=i, column=6, value=stats["tokenValue"])
        ws.cell(row=i, column=7, value=stats["cost"])
        ws.cell(row=i, column=7).number_format = '#,##0.000000'
        ws.cell(row=i, column=8, value=stats["count"])

    # Totals row
    total_row = len(sorted_users) + 2
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=sum(s["inputTokens"] for _, s in sorted_users))
    ws.cell(row=total_row, column=4, value=sum(s["writeTokens"] for _, s in sorted_users))
    ws.cell(row=total_row, column=5, value=sum(s["readTokens"] for _, s in sorted_users))
    ws.cell(row=total_row, column=6, value=sum(s["tokenValue"] for _, s in sorted_users))
    ws.cell(row=total_row, column=7, value=sum(s["cost"] for _, s in sorted_users))
    ws.cell(row=total_row, column=7).number_format = '#,##0.000000'
    ws.cell(row=total_row, column=8, value=sum(s["count"] for _, s in sorted_users))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def main():
    args = parse_args()

    start = parse_date(args.start)
    end = parse_date(args.end)

    if start > end:
        sys.exit("Error: start date must be <= end date")

    print(f"Connecting to MongoDB...")
    client = MongoClient(args.connection)
    db = client.get_default_database()
    if db is None:
        sys.exit("Error: connection string must include a database name")

    print(f"Fetching users...")
    users = fetch_users(db)
    print(f"  Found {len(users)} users")

    print(f"Fetching transactions from {args.start} to {args.end}...")
    transactions = fetch_transactions(db, start, end)
    print(f"  Found {len(transactions)} transaction records")

    if not transactions:
        print("No transactions found in the given date range.")
        return

    wb = Workbook()
    write_raw_sheet(wb, transactions, users)
    write_summary_sheet(wb, transactions, users)

    wb.save(args.output)
    print(f"Exported to {args.output}")

    # Print quick summary
    total_cost = sum(compute_token_cost(t) for t in transactions)
    print(f"\n--- Summary ---")
    print(f"Date range:  {args.start} ~ {args.end}")
    print(f"Records:     {len(transactions)}")
    print(f"Users:       {len(set(str(t.get('user', '')) for t in transactions))}")
    print(f"Total cost:  ${total_cost:,.6f} USD")

    client.close()


if __name__ == "__main__":
    main()